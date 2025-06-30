from openpyxl import load_workbook
from Constantes import constantes_historial
from Constantes import constantes_items
class Facade_Historial():
    def __init__(self):
        self

    def obtener_historial(self, path_excel):
        excel_dataframe=load_workbook(path_excel, data_only=True)
        historial= excel_dataframe["Historial"]
        datos_totales = []
        datos_actual = []
        row = 2
        while historial.cell(row=row, column=1).value:
            columna = 1
            while(columna < 41):
                datos_actual.append(historial.cell(row=row, column=columna).value)
                columna += 1
            datos_totales.append(datos_actual)
            datos_actual = []
            row += 1

        return datos_totales
    
    #Escribe el historial 
    def escribir_historial(self, datos_historial, historial_hoja, row):
        cant_var = 1
        i = 0
        while i < constantes_historial.limite_hasta_items:
            historial_hoja.cell(row=row, column=cant_var, value=datos_historial[i])
            cant_var += 1
            i += 1

        cant_var_ant = cant_var
        row_ant = row
        j = 0
        k = 0
        productos_servicios = datos_historial[constantes_historial.limite_hasta_items]
        while j < len(productos_servicios):
            k = 0
            cant_var = cant_var_ant
            while k < constantes_historial.cant_datos_items:
                historial_hoja.cell(row=row, column=cant_var, value=productos_servicios[j][k])
                k += 1
                cant_var += 1
            historial_hoja.cell(row=row, column=constantes_historial.columna_id_excel, value=datos_historial[0])
            row += 1
            j += 1

        i += 1
        row = row_ant

        while i < constantes_historial.limite_hasta_tributos:
            historial_hoja.cell(row=row, column=cant_var, value=datos_historial[i])
            cant_var += 1
            i += 1

        cant_var_ant = cant_var
        row_ant = row
        
        tributos = datos_historial[constantes_historial.limite_hasta_tributos]
        for producto_actual in productos_servicios:
            j = 0
            tributo_found = False
            while j < len(tributos) and tributo_found == False:
                if(tributos[j][constantes_items.pos_id_relacionado_tributo] == producto_actual[constantes_items.pos_producto_servicio]):
                    historial_hoja.cell(row=row, column=cant_var, value=tributos[j][constantes_items.pos_descripcion_impuesto_adicional])
                    cant_var +=1
                    historial_hoja.cell(row=row, column=cant_var, value=tributos[j][constantes_items.pos_alicuota_impuesto_adicional])
                    row+=1
                    tributo_found = True
                j+=1
            cant_var = cant_var_ant

        row = row_ant
        cant_var = cant_var_ant + constantes_historial.cant_filas_a_saltear
        i+=1

        while i < constantes_historial.numero_finalizacion:
            historial_hoja.cell(row=row, column=cant_var, value=datos_historial[i])
            cant_var += 1
            i += 1

    def obtener_row_historial(self, dataframe_historial):
        row = 2
        while dataframe_historial.cell(row=row, column=constantes_historial.columna_id_excel).value or dataframe_historial.cell(row=row, column=constantes_historial.columna_item).value:
            row += 1
        
        return row

    def guardar_historial(self, dataframe_historial, file_name):
        dataframe_historial.save(file_name)