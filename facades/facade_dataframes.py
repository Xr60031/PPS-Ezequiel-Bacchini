from openpyxl import load_workbook

class Facade_dataframe():
    def __init__(self):
        self

    def obtener_dataframe_historial(self, file_name):
        excel_dataframe=load_workbook(file_name, data_only=True)
        datos_historial = []
        datos_historial.append(excel_dataframe)
        datos_historial.append(excel_dataframe["Historial"])
        return datos_historial
    
    def obtener_dataframe_facturar(self, file_name):
        excel_dataframe=load_workbook(file_name, data_only=True)
        return excel_dataframe["Facturas_A_Realizar"]
    
    def obtener_xml(self, file_name):
        excel_dataframe=load_workbook(file_name, data_only=True)
        workbook = excel_dataframe["XML"]
        xml = workbook.cell(row=1, column=1).value
        excel_dataframe.save(file_name)
        return xml
    
    def guardar_xml(self, file_name, xml):
        excel_dataframe=load_workbook(file_name, data_only=True)
        workbook = excel_dataframe["XML"]
        workbook.cell(row=1, column=1, value=xml)
        excel_dataframe.save(file_name)
