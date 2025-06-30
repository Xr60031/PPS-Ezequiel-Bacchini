from openpyxl import load_workbook

class Facade_Configuracion():
    def __init__(self):
        self
    
    def guardar_configuracion_inicial(self, template_path, config, copy_path):
        excel_dataframe=load_workbook(template_path, data_only=False)
        datos_hoja=excel_dataframe["Datos"]
        datos_hoja["A2"]=config[0]
        datos_hoja["B2"]=config[1]
        datos_hoja["C2"]=config[2]
        datos_hoja["D2"]=config[3]
        datos_hoja["E2"]=config[4]
        datos_hoja["F2"]=config[5]
        datos_hoja["G2"]=config[6]
        datos_hoja["H2"]=config[7]
        datos_hoja["I2"]=config[8]
        datos_hoja["J2"]=config[9]
        datos_hoja["K2"]=config[10]
        datos_hoja["L2"]=config[11]
        excel_dataframe.save(copy_path)
    
    def guardar_configuracion(self, path_excel, config):

        excel_dataframe=load_workbook(path_excel)
        
        datos_hoja=excel_dataframe["Datos"]
        
        if config[0]:
            datos_hoja["A2"]=config[0]
        
        if config[1]:
            datos_hoja["B2"]=config[1]
        
        if config[2]:
            datos_hoja["C2"]=config[2]
        
        if config[3]:
            datos_hoja["D2"]=config[3]
        
        if config[4]:
            datos_hoja["E2"]=config[4]
        
        if config[5]:
            datos_hoja["F2"]=config[5]
        
        if config[6]:
            datos_hoja["G2"]=config[6]
        
        if config[7]:
            datos_hoja["H2"]=config[7]
        
        if config[8]:
            datos_hoja["I2"]=config[8]
        
        if config[9]:
            datos_hoja["J2"]=config[9]
        
        if config[10]:
            datos_hoja["K2"]=config[10]
        
        if config[11]:
            datos_hoja["L2"]=config[11]
        
        excel_dataframe.save(path_excel)