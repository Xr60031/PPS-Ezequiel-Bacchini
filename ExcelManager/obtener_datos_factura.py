from openpyxl import load_workbook
from ExcelManager.obtenedor_datos_excel import Obtenedor_Datos_Excel
from Funciones.Ensambladores.armar_item import Armador_Item

from Constantes.Excel.constantes_excel import constantes_genericas_excel, workbooks
from Constantes.Facturacion.constantes_arrays import constantes_array_datos_factura

class Obtenedor_Datos_Factura():
    def __init__(self):
        self

    def obtener_ID(self, factura_hoja, row):
        return factura_hoja.cell(row=row, column=constantes_genericas_excel.starting_colvar.value).value

    def obtener_productos_servicios(self, excel_dataframe, datos_factura, row, productos_servicios, tributos, worksheet): 
        ID = datos_factura[constantes_array_datos_factura.pos_ID.value]
        ID_anterior = ID
        hoja_actual = excel_dataframe[worksheet]

        while ID == ID_anterior and ID != "" :
            obtenedor_datos_excel = Obtenedor_Datos_Excel()
            datos_hoja_factura = obtenedor_datos_excel.obtener_datos_hoja_factura_item_actual(hoja_actual, row)
            hoja_actual = excel_dataframe[workbooks.items.value]
            datos_hoja_item = obtenedor_datos_excel.obtener_datos_hoja_item_actual(hoja_actual, datos_hoja_factura)
            armador_item = Armador_Item()
            armador_item.armar_datos_producto_servicio(productos_servicios, tributos, datos_hoja_factura, datos_hoja_item)
            
            ID_anterior = ID
            row += 1
            hoja_actual = excel_dataframe[worksheet]
            ID = self.obtener_ID(hoja_actual, row) 

    #Abre la plantilla de excel y devuelve una lista con todos los datos que el sistema necesita para facturar
    def obtener_datos_factura(self, filename, worksheet):
        excel_dataframe=load_workbook(filename, data_only=True)
        factura_hoja=excel_dataframe[worksheet]
        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value
        datos_facturas_total = []

        ID = factura_hoja.cell(row=row, column=cant_var).value
        ID_anterior = None
        while ID:
            datos_factura = []
            if ID != ID_anterior:
                while cant_var < constantes_array_datos_factura.pos_ultimo_dato_antes_de_productos.value:
                    datos_factura.append(factura_hoja.cell(row=row, column=cant_var).value)
                    cant_var+=1
                
                productos_servicios = []
                tributos = []
                self.obtener_productos_servicios(excel_dataframe, datos_factura, row, productos_servicios, tributos, worksheet)
                datos_factura.append(tributos)
                datos_factura.append(productos_servicios)
                factura_hoja=excel_dataframe[worksheet]
                ID_anterior = ID
            if(datos_factura != []):
                datos_facturas_total.append(datos_factura)
            cant_var = constantes_genericas_excel.starting_colvar.value
            row += 1
            ID = factura_hoja.cell(row=row, column=constantes_genericas_excel.starting_colvar.value).value

        return datos_facturas_total