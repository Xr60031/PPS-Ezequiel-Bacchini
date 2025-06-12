from openpyxl import load_workbook

def agregar_producto_servicio(path_excel, datos):
  excel_dataframe=load_workbook(path_excel, data_only = True)
  items=excel_dataframe["Items"]

  row = items.max_row + 1
  column = 1

  i = 0
  while i < 7:
    if datos[i] is not None:
      items.cell(row=row, column=column, value=datos[i])
    column += 1
    i += 1

  excel_dataframe.save(path_excel)

def modificar_producto_servicio(path_excel, datos, nombre_item_target):
  excel_dataframe=load_workbook(path_excel, data_only = True)
  items=excel_dataframe["Items"]

  row = 2
  cant_var = 1
  nombre_a_buscar = nombre_item_target
  nombre= items.cell(row=row, column=cant_var).value

  while nombre != nombre_a_buscar:
    row +=1
    nombre= items.cell(row=row, column=cant_var).value
  
  num_dato = 0
  while num_dato < 7: 
    if(datos[num_dato]!= None and datos[num_dato] != ""):
      items.cell(row=row, column=cant_var, value=datos[num_dato])
    cant_var += 1
    num_dato += 1

  excel_dataframe.save(path_excel)
  
def eliminar_producto_servicio(path_excel, nombre_item_target):
  excel_dataframe=load_workbook(path_excel, data_only = True)
  items=excel_dataframe["Items"]

  row = 2
  cant_var = 1
  nombre_a_buscar = nombre_item_target
  nombre = items.cell(row=row, column=cant_var).value
  while nombre != nombre_a_buscar:
    row +=1
    nombre= items.cell(row=row, column=cant_var).value
  
  items.cell(row=row, column=cant_var, value="")
  cant_var += 1
  items.cell(row=row, column=cant_var, value="")
  cant_var += 1
  items.cell(row=row, column=cant_var, value="")
  cant_var += 1
  items.cell(row=row, column=cant_var, value="")
  cant_var += 1
  items.cell(row=row, column=cant_var, value="")
  cant_var += 1
  items.cell(row=row, column=cant_var, value="")
  cant_var += 1
  items.cell(row=row, column=cant_var, value="")

  #Reacomodo las filas
  row += 1
  cant_var = 1
  while row <= items.max_row and items.cell(row, column=1).value:
    while cant_var < 8:
      dato = items.cell(row=row, column=cant_var).value
      items.cell(row=row-1, column=cant_var, value=dato)
      items.cell(row=row, column=cant_var, value="")
      cant_var += 1
    cant_var=1
    row +=1
    
  excel_dataframe.save(path_excel)

def obtener_productos_servicios(path_excel_):
  excel_dataframe=load_workbook(path_excel_, data_only = True)
  items=excel_dataframe["Items"]

  items_total = []
  items_actuales = []

  row = 2
  cant_var = 1

  nombre= items.cell(row=row, column=cant_var).value
  while nombre is not None:
    cant_var +=1
    codigo = items.cell(row=row, column=cant_var).value

    cant_var +=1
    descripcion_prod_ser = items.cell(row=row, column=cant_var).value

    cant_var +=1
    precio_unitario = items.cell(row=row, column=cant_var).value

    cant_var +=1
    impuesto_adicional = items.cell(row=row, column=cant_var).value

    cant_var +=1
    descripcion_ia = items.cell(row=row, column=cant_var).value

    cant_var +=1
    alicuota = items.cell(row=row, column=cant_var).value

    items_actuales.append(nombre)
    items_actuales.append(codigo)
    items_actuales.append(descripcion_prod_ser)
    items_actuales.append(float(precio_unitario))
    items_actuales.append(impuesto_adicional)
    items_actuales.append(descripcion_ia)
    items_actuales.append(alicuota)

    items_total.append(items_actuales)
    items_actuales = []
    row +=1

    cant_var = 1
    nombre= items.cell(row=row, column=cant_var).value
  
  excel_dataframe.close()
  return items_total