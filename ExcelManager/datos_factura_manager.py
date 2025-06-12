from openpyxl import load_workbook
from funciones.obtener_ID_tributo import obtener_ID_tributo
from datetime import datetime

def obtener_ID_documento(documento):
    return{
        'CUIT' : 80,
        'CDI' : 87,
        'CI EXTRANJERA' : 91,
        'PASAPORTE' : 94,
        'DNI' : 96,
        'OTRO' : 99
    }.get(documento)

def obtener_ID_concepto(concepto):
    return{
        'PRODUCTOS' : 1,
        'SERVICIOS' : 2,
        'PRODUCTOS Y SERVICIOS' : 3
    }.get(concepto)

def obtener_ID_Factura(tipo_factura):
    return{
        'Factura C' : 11,
    }.get(tipo_factura)

def obtener_ID_condicion_iva_cliente(tipo_iva):
    return{
        'Consumidor Final': 5
    }.get(tipo_iva)

#Arma la biblioteca con los datos de la factura
def armar_biblioteca_factura(datos_factura):
    datos_factura = {
        "Identificador_Factura": datos_factura[0],
        "tipo_factura_nota": datos_factura[1],
        "Nombre_y_Apellido_Cliente": datos_factura[2],
        "Tipo_Documento": datos_factura[3],
        "Numero_de_documento_del_cliente": int(datos_factura[4]),
        "Condicion_de_venta_Cliente": datos_factura[5],
        "Condicion_frente_al_IVA_Cliente": datos_factura[6],
        "Concepto": datos_factura[7],
        "Fecha_servicio_desde": datos_factura[8],
        "Fecha_servicio_hasta": datos_factura[9],
        "Fecha_vencimiento_de_pago": datos_factura[10],
        "ID_doc": obtener_ID_documento(datos_factura[3]),
        "ID_concepto": obtener_ID_concepto(datos_factura[7]),
        "ID_factura_nota": obtener_ID_Factura(datos_factura[1]),
        "ID_IVA_cliente": obtener_ID_condicion_iva_cliente(datos_factura[6]),
        "Importe_Neto": 0,
        "Importe_Total": 0,
        "Importe_Tributo": 0,
        "tributos": 0,
        "items": 0,
        "nro_cbte_anular" : None,
        "fecha_emision" : datetime.now()
    }
    return datos_factura

def set_comprobante_anular(datos_factura, cbte_anular):
    datos_factura["nro_cbte_anular"] = cbte_anular


def obtener_datos_hoja_factura_item_actual(factura_hoja, row):
    cantidad = float(factura_hoja.cell(row=row, column=14).value or 0)
    porcentaje_bonificado = factura_hoja.cell(row=row, column=13).value
    nombre_producto_servicio = factura_hoja.cell(row=row, column=12).value

    datos_hoja_factura = [
        int(cantidad),
        porcentaje_bonificado or 0,
        nombre_producto_servicio
    ]
    return datos_hoja_factura

def obtener_datos_hoja_item_actual(items, datos_hoja_factura):
    row_ = 2
    while items.cell(row=row_, column=1).value and datos_hoja_factura[2] != items.cell(row=row_, column=1).value:
        row_ += 1

    codigo_producto = items.cell(row=row_, column=2).value
    descripcion = items.cell(row=row_, column=3).value
    precio_unitario = float(items.cell(row=row_, column=4).value or 0)
    impuesto_adicional = items.cell(row=row_, column=5).value
    descripcion_tributo_adicional = items.cell(row=row_, column=6).value
    alicuota = float(items.cell(row=row_, column=7).value or 0)
    
    importe_bonificado = datos_hoja_factura[0]*precio_unitario*(100-datos_hoja_factura[1])/100
    subtotal = importe_bonificado*(100-alicuota)/100

    datos_hoja_item = [
        codigo_producto,
        descripcion,
        precio_unitario,
        impuesto_adicional,
        descripcion_tributo_adicional,
        alicuota,
        importe_bonificado,
        subtotal
    ]

    return datos_hoja_item

def armar_datos_producto_servicio(productos_servicios, tributos, datos_hoja_factura, datos_hoja_item):
    productos_y_servicios_actual = [
        datos_hoja_factura[2],  # Producto/Servicio
        datos_hoja_factura[1],  # Porcentaje Bonificado
        datos_hoja_factura[0],  # Cantidad
        datos_hoja_item[0],  # Código Producto
        datos_hoja_item[1],  # Descripción
        datos_hoja_item[2],  # Precio Unitario
        datos_hoja_item[3],  # Impuesto Adicional
        datos_hoja_item[6],  # Importe Bonificado
        datos_hoja_item[7]# Subtotal
    ] 

    productos_servicios.append(productos_y_servicios_actual)

    id_tributo = obtener_ID_tributo(datos_hoja_item[3])
    if(id_tributo):

        tributos_actual = [
            id_tributo, #ID_tributo
            datos_hoja_item[4], #descripcion_impuesto_adicional
            datos_hoja_item[6], #Precio_Bonificado 
            datos_hoja_item[5], #alicuota_impuesto_adicional
            datos_hoja_item[6]*datos_hoja_item[5]#Precio_Bonificado*alicuota_impuesto_adicional
        ]
        tributos.append(tributos_actual)

    return

def obtener_ID(factura_hoja, row):
    return factura_hoja.cell(row=row, column=1).value

def obtener_productos_servicios(excel_dataframe, datos_factura, biblioteca_datos_vendedor, row): 
    ID = datos_factura["Identificador_Factura"]
    ID_anterior = ID
    i=0
    imp_neto=0
    imp_total=0
    imp_total = float(imp_total)
    imp_tributo=0
    productos_servicios = []
    tributos = []
    hoja_actual = excel_dataframe["Facturas_A_Realizar"]

    while ID == ID_anterior and ID != "" :
        datos_hoja_factura = obtener_datos_hoja_factura_item_actual(hoja_actual, row)
        hoja_actual = excel_dataframe["Items"]
        datos_hoja_item = obtener_datos_hoja_item_actual(hoja_actual, datos_hoja_factura)
        armar_datos_producto_servicio(productos_servicios, tributos, datos_hoja_factura, datos_hoja_item)
        # + subtotal
        imp_total += productos_servicios[i][8]
        # + importe_bonificado
        imp_neto += productos_servicios[i][7]
        
        ID_anterior = ID
        i += 1
        row += 1
        hoja_actual = excel_dataframe["Facturas_A_Realizar"]
        ID = obtener_ID(hoja_actual, row) 

    if biblioteca_datos_vendedor["id_tributo_global"]:
        tributos.append([
            #ID_tributo
            biblioteca_datos_vendedor["id_tributo_global"],
            #descripcion_impuesto_adicional
            biblioteca_datos_vendedor["desc_iag"],
            #imp_neto
            imp_neto,
            #alicuota_impuesto_adicional
            biblioteca_datos_vendedor["alicuota"],
            #imp_neto*alicuota_impuesto_adicional
            imp_neto*(biblioteca_datos_vendedor["alicuota"] or 0)/100
        ])
        imp_total += imp_neto*(biblioteca_datos_vendedor["alicuota"] or 0)/100

    datos_factura["tributos"] = tributos
    if len(datos_factura["tributos"]) == 0:
        datos_factura["Importe_Tributo"] = 0
    else:
        imp_tributo = imp_total - imp_neto
        datos_factura["Importe_Tributo"] = imp_tributo

    datos_factura["Importe_Total"] = imp_total
    datos_factura["Importe_Neto"] = imp_neto
    datos_factura["tributos"] = tributos
    datos_factura["items"] = productos_servicios

#Abre la plantilla de excel y devuelve una lista con todos los datos del cliente que el sistema necesita
def obtener_datos_factura(filename, biblioteca_datos_vendedor):
    excel_dataframe=load_workbook(filename, data_only=True)
    factura_hoja=excel_dataframe["Facturas_A_Realizar"]
    row = 2
    cant_var = 1
    datos_facturas_total = []

    ID = factura_hoja.cell(row=row, column=cant_var).value
    ID_anterior = None
    while ID:
        datos_factura = []
        if ID != ID_anterior:
            #ID
            #tipo_factura_nota
            #nombre_apellido_cliente
            #tipo_documento
            #nro_doc_cliente
            #condicion_venta_cliente
            #condicion_IVA_cliente
            #concepto_producto_servicio
            #fecha_ser_desde
            #fecha_ser_hasta
            #fecha_vencimiento_pago
            while cant_var < 12:
                datos_factura.append(factura_hoja.cell(row=row, column=cant_var).value)
                cant_var+=1

            
            biblioteca_factura = armar_biblioteca_factura(datos_factura)
            obtener_productos_servicios(excel_dataframe, biblioteca_factura, biblioteca_datos_vendedor, row)
            factura_hoja=excel_dataframe["Facturas_A_Realizar"]
            ID_anterior = ID
        datos_facturas_total.append(biblioteca_factura)
        cant_var = 1
        row += 1
        ID = factura_hoja.cell(row=row, column=1).value

    return datos_facturas_total