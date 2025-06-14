from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from Constantes.constantes_columna import col_hoja_facturacion

def limpiar_hoja_facturacion(path):
        wb = load_workbook(path, data_only=False)
        workbook = wb["Facturas_A_Realizar"]
        column_start = col_hoja_facturacion

        for j in range(column_start, workbook.max_column + 1):
            col_letter = get_column_letter(j)
            if col_letter in workbook.column_dimensions:
                del workbook.column_dimensions[col_letter]

        for row in workbook.iter_rows(min_row=2, max_row=workbook.max_row, min_col=1, max_col=workbook.max_column):
            workbook.row_dimensions[0].height = None
            for cell in row:
                cell._style = None
                cell.value = None
        
        wb.save(path)
