from openpyxl import load_workbook

#Abre la plantilla de excel y devuelve una lista con todos los datos del usuario que el sistema necesita
def obtener_datos_usuario(file_name):
    excel_dataframe=load_workbook(file_name, data_only=True)
    datos_basicos_hoja=excel_dataframe["Datos"]
    info_usuario = []

    datos_usuario=1

    #nombre
    #cuit
    #nombre empresa
    #punto venta
    #razon social
    #domicilio
    #condicion iva vendedor
    #ingresos brutos
    #fecha inicio
    #impuesto adicional global
    #descripcion iag
    #alicuota
    while datos_usuario < 13:
        info_usuario.append(datos_basicos_hoja.cell(row=2, column=datos_usuario).value)
        datos_usuario+=1

    excel_dataframe.close()

    return info_usuario