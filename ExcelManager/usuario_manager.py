from openpyxl import load_workbook
from funciones.obtener_ID_tributo import obtener_ID_tributo

#Abre la plantilla de excel y devuelve una lista con todos los datos del usuario que el sistema necesita
def obtener_datos_usuario(file_name):
    excel_dataframe=load_workbook(file_name, data_only=True)
    datos_basicos_hoja=excel_dataframe["Datos"]
    info_usuario = []

    datos_usuario=1

    #nombre
    #cuit
    #nombre empresa
    #punto venta
    #razon social
    #domicilio
    #condicion iva vendedor
    #ingresos brutos
    #fecha inicio
    #impuesto adicional global
    #descripcion iag
    #alicuota
    while datos_usuario < 13:
        info_usuario.append(datos_basicos_hoja.cell(row=2, column=datos_usuario).value)
        datos_usuario+=1

    excel_dataframe.close()

    return info_usuario

#Arma la biblioteca con los datos del vendedor
def armar_biblioteca_vendedor(datos_vendedor):
    bibloteca_datos_vendedor = {
        "Nombre": datos_vendedor[0],
        "CUIT": datos_vendedor[1],
        "Nombre_Empresa": datos_vendedor[2],
        "Punto_de_venta": datos_vendedor[3],
        "Razon_Social": datos_vendedor[4],
        "Domicilio": datos_vendedor[5],
        "Condicion_frente_al_IVA": datos_vendedor[6],
        "Ingresos_Brutos": datos_vendedor[7],
        "Fecha_Inicio": datos_vendedor[8],
        "iag": datos_vendedor[9],
        "desc_iag": datos_vendedor[10],
        "alicuota": datos_vendedor[11],
        "id_tributo_global": obtener_ID_tributo(datos_vendedor[9])
    }

    return bibloteca_datos_vendedor