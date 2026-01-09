from openpyxl import load_workbook
from Constantes.Facturacion.constantes_items import constantes_posicion_items, constantes_posicion_tributos
from Constantes.Excel.constantes_excel import constantes_genericas_excel, workbooks
from Constantes.Facturacion.constantes_arrays import constantes_armador_historial
class Facade_Historial():
    def __init__(self):
        self

    def obtener_historial(self, path_excel):
        excel_dataframe=load_workbook(path_excel, data_only=True)
        historial= excel_dataframe[workbooks.historial.value]
        datos_totales = []
        datos_actual = []
        row = constantes_genericas_excel.starting_row.value
        while historial.cell(row=row, column=constantes_genericas_excel.starting_colvar.value).value:
            columna = constantes_genericas_excel.starting_colvar.value
            while(columna < constantes_armador_historial.cantidad_columnas_datos.value):
                datos_actual.append(historial.cell(row=row, column=columna).value)
                columna += 1
            datos_totales.append(datos_actual)
            datos_actual = []
            row += 1

        return datos_totales
    
    #Escribe el historial 
    def escribir_historial(self, datos_historial, historial_hoja, row):
        columna_numero = constantes_genericas_excel.starting_colvar.value
        productos_servicios = None
        for dato in datos_historial:
            #-1 Porque las columnas en excel comienzan en 1 y las listas comienzan en 0
            #Escritura de Productos/Servicios
            if columna_numero-1 == constantes_armador_historial.limite_hasta_items.value:
                productos_servicios = dato
                numero_columna_aux = columna_numero
                row_ant = row
                for item in dato:
                    for dato_item in range(0, constantes_armador_historial.cant_datos_items.value):
                        historial_hoja.cell(row=row, column=numero_columna_aux, value=item[dato_item])
                        numero_columna_aux += 1
                    row += 1
                    numero_columna_aux = columna_numero
                    
                columna_numero += constantes_armador_historial.cant_datos_items.value
                row = row_ant
            elif columna_numero == constantes_armador_historial.limite_hasta_tributos.value+constantes_armador_historial.distancia_entre_tributos_items.value and dato:
                print("ingreso tributos", dato)
                numero_columna_aux = columna_numero
                row_ant = row
                for item in productos_servicios:
                    tributo_found = False
                    i = 0
                    while i < len(dato) and tributo_found == False:
                        if dato[i][constantes_posicion_tributos.pos_id_relacionado_tributo.value] == item[constantes_posicion_items.pos_producto_servicio.value]:
                            historial_hoja.cell(row=row, column=numero_columna_aux, value=dato[i][constantes_posicion_tributos.pos_descripcion_impuesto_adicional.value])
                            numero_columna_aux +=1
                            historial_hoja.cell(row=row, column=numero_columna_aux, value=dato[i][constantes_posicion_tributos.pos_alicuota_impuesto_adicional.value])
                            tributo_found = True
                        i+=1
                    row += 1
                    numero_columna_aux = columna_numero
                    
                columna_numero += constantes_armador_historial.cant_datos_tributos_utilizados.value
                row = row_ant
            elif columna_numero == constantes_armador_historial.limite_hasta_tributos.value+constantes_armador_historial.distancia_entre_tributos_items.value:
                columna_numero += constantes_armador_historial.cant_datos_tributos_utilizados.value
            else:
                historial_hoja.cell(row=row, column=columna_numero, value=dato)
                columna_numero+=1

    def obtener_row_historial(self, dataframe_historial):
        row = constantes_genericas_excel.starting_colvar.value
        while dataframe_historial.cell(row=row, column=constantes_armador_historial.columna_id_excel.value).value or dataframe_historial.cell(row=row, column=constantes_armador_historial.columna_item.value).value:
            row += 1
        
        return row

    def guardar_historial(self, dataframe_historial, file_name):
        dataframe_historial.save(file_name)

    def get_ultimo_ID_factura_usado(self, path_excel):
        excel_dataframe=load_workbook(path_excel, data_only=True)
        historial= excel_dataframe[workbooks.historial.value]
        row = constantes_genericas_excel.starting_row.value
        ultimo_ID = None
        while historial.cell(row=row, column=constantes_genericas_excel.starting_colvar.value).value:
            ultimo_ID = historial.cell(row=row, column=constantes_genericas_excel.starting_colvar.value).value
            row+=1
        return ultimo_ID