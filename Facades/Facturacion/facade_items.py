from openpyxl import load_workbook
from Constantes.Excel.constantes_excel import constantes_genericas_excel, workbooks
from Constantes.Facturacion.constantes_items import constantes_posicion_items
class Facade_Items():
    def __init__(self):
        self

    def obtener_productos_servicios(self, path_excel_):
        excel_dataframe=load_workbook(path_excel_, data_only = True)
        items=excel_dataframe[workbooks.items.value]

        items_total = []
        items_actuales = []

        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value

        nombre= items.cell(row=row, column=cant_var).value
        while nombre is not None:
            cant_var +=1
            codigo = items.cell(row=row, column=cant_var).value

            cant_var +=1
            descripcion_prod_ser = items.cell(row=row, column=cant_var).value

            cant_var +=1
            precio_unitario = items.cell(row=row, column=cant_var).value

            cant_var +=1
            impuesto_adicional = items.cell(row=row, column=cant_var).value

            cant_var +=1
            descripcion_ia = items.cell(row=row, column=cant_var).value

            cant_var +=1
            alicuota = items.cell(row=row, column=cant_var).value

            items_actuales.append(nombre)
            items_actuales.append(codigo)
            items_actuales.append(descripcion_prod_ser)
            items_actuales.append(float(precio_unitario))
            items_actuales.append(impuesto_adicional)
            items_actuales.append(descripcion_ia)
            items_actuales.append(alicuota)

            items_total.append(items_actuales)
            items_actuales = []
            row +=1

            cant_var = constantes_genericas_excel.starting_colvar.value
            nombre= items.cell(row=row, column=cant_var).value
        
        excel_dataframe.close()
        return items_total

    def eliminar_producto_servicio(self, path_excel, nombre_item_target):
        excel_dataframe=load_workbook(path_excel, data_only = True)
        items=excel_dataframe[workbooks.items.value]

        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value
        nombre_a_buscar = nombre_item_target
        nombre = items.cell(row=row, column=cant_var).value
        while nombre != nombre_a_buscar:
            row +=1
            nombre= items.cell(row=row, column=cant_var).value
        

        i = 0
        while i < constantes_posicion_items.cantidad_total_datos_por_item.value:
            items.cell(row=row, column=cant_var, value="")
            cant_var += 1
            i+=1

        #Reacomodo las filas
        row += 1
        cant_var = constantes_genericas_excel.starting_colvar.value
        while row <= items.max_row and items.cell(row, column=constantes_genericas_excel.starting_colvar.value).value:
            while cant_var < constantes_posicion_items.cantidad_total_datos_por_item.value:
                dato = items.cell(row=row, column=cant_var).value
                items.cell(row=row-1, column=cant_var, value=dato)
                items.cell(row=row, column=cant_var, value="")
                cant_var += 1
            cant_var= constantes_genericas_excel.starting_colvar.value
            row +=1
            
        excel_dataframe.save(path_excel)

    def agregar_producto_servicio(self, path_excel, datos):
        excel_dataframe=load_workbook(path_excel, data_only = True)
        items=excel_dataframe[workbooks.items.value]

        row = items.max_row + 1
        column = constantes_genericas_excel.starting_colvar.value

        i = 0
        while i < constantes_posicion_items.cantidad_total_datos_por_item.value-2:
            if datos[i] is not None:
                items.cell(row=row, column=column, value=datos[i])
            column += 1
            i += 1

        excel_dataframe.save(path_excel)

    def modificar_producto_servicio(self, path_excel, datos, nombre_item_target):
        excel_dataframe=load_workbook(path_excel, data_only = True)
        items=excel_dataframe[workbooks.items.value]

        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value
        nombre_a_buscar = nombre_item_target
        nombre= items.cell(row=row, column=cant_var).value

        while nombre != nombre_a_buscar:
            row +=1
            nombre= items.cell(row=row, column=cant_var).value
        
        num_dato = 0
        while num_dato < constantes_posicion_items.cantidad_total_datos_por_item.value-2: 
            if(datos[num_dato]!= None and datos[num_dato] != ""):
                items.cell(row=row, column=cant_var, value=datos[num_dato])
            cant_var += 1
            num_dato += 1

        excel_dataframe.save(path_excel)