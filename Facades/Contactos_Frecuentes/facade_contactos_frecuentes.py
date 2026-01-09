from openpyxl import load_workbook
from Constantes.Excel.constantes_excel import constantes_genericas_excel, workbooks, constantes_clientes_frecuentes_excel
class Facade_Clientes():
    def __init__(self):
        self

    def obtener_clientes(self, path_excel_):
        excel_dataframe=load_workbook(path_excel_, data_only = True)
        clientes=excel_dataframe[workbooks.contactos_frecuentes.value]

        clientes_total = []
        clientes_actuales = []

        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value

        nombre_apellido_cliente= clientes.cell(row=row, column=cant_var).value
        while nombre_apellido_cliente is not None:
            cant_var +=1
            nro_doc_cliente = clientes.cell(row=row, column=cant_var).value

            cant_var +=1
            nro_telefono_cliente = clientes.cell(row=row, column=cant_var).value

            cant_var +=1
            provincia_cliente = clientes.cell(row=row, column=cant_var).value

            cant_var +=1
            localidad_cliente = clientes.cell(row=row, column=cant_var).value

            cant_var +=1
            domicilio_cliente = clientes.cell(row=row, column=cant_var).value
            
            cant_var +=1
            condicion_venta_cliente = clientes.cell(row=row, column=cant_var).value
            
            cant_var +=1
            condicion_iva_cliente = clientes.cell(row=row, column=cant_var).value

            clientes_actuales.append(nombre_apellido_cliente)
            clientes_actuales.append(nro_doc_cliente)
            clientes_actuales.append(nro_telefono_cliente)
            clientes_actuales.append(provincia_cliente)
            clientes_actuales.append(localidad_cliente)
            clientes_actuales.append(domicilio_cliente)
            clientes_actuales.append(condicion_venta_cliente)
            clientes_actuales.append(condicion_iva_cliente)

            clientes_total.append(clientes_actuales)
            clientes_actuales = []
            row +=1

            cant_var = constantes_genericas_excel.starting_colvar.value
            nombre_apellido_cliente= clientes.cell(row=row, column=cant_var).value
        
        excel_dataframe.close()
        return clientes_total

    def eliminar_cliente(self, path_excel, nombre_cliente_target):
        excel_dataframe=load_workbook(path_excel, data_only = True)
        clientes=excel_dataframe[workbooks.contactos_frecuentes.value]

        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value
        nombre_a_buscar = nombre_cliente_target
        nombre = clientes.cell(row=row, column=cant_var).value
        while nombre != nombre_a_buscar:
            row +=1
            nombre= clientes.cell(row=row, column=cant_var).value
        
        i = 0
        while i < constantes_clientes_frecuentes_excel.cantidad_total_variables_cliente.value:
            clientes.cell(row=row, column=cant_var, value="")
            cant_var += 1
            i+=1

        #Reacomodo las filas
        row += 1
        cant_var = constantes_genericas_excel.starting_colvar.value
        while row <= clientes.max_row and clientes.cell(row, column=constantes_genericas_excel.starting_colvar.value).value:
            while cant_var < constantes_clientes_frecuentes_excel.cantidad_total_variables_cliente.value:
                dato = clientes.cell(row=row, column=cant_var).value
                clientes.cell(row=row-1, column=cant_var, value=dato)
                clientes.cell(row=row, column=cant_var, value="")
                cant_var += 1
            cant_var= constantes_genericas_excel.starting_colvar.value
            row +=1
            
        excel_dataframe.save(path_excel)

    def agregar_cliente(self, path_excel, datos):
        excel_dataframe=load_workbook(path_excel, data_only = True)
        clientes=excel_dataframe[workbooks.contactos_frecuentes.value]

        row = clientes.max_row + 1
        column = constantes_genericas_excel.starting_colvar.value

        print(datos)
        i = 0
        while i < constantes_clientes_frecuentes_excel.cantidad_total_variables_cliente.value-1:
            if datos[i] is not None:
                clientes.cell(row=row, column=column, value=datos[i])
            column += 1
            i += 1

        excel_dataframe.save(path_excel)

    def modificar_cliente(self, path_excel, datos, nombre_item_target):
        excel_dataframe=load_workbook(path_excel, data_only = True)
        clientes=excel_dataframe[workbooks.contactos_frecuentes.value]

        row = constantes_genericas_excel.starting_row.value
        cant_var = constantes_genericas_excel.starting_colvar.value
        nombre_a_buscar = nombre_item_target
        nombre= clientes.cell(row=row, column=cant_var).value

        while nombre != nombre_a_buscar:
            row +=1
            nombre= clientes.cell(row=row, column=cant_var).value
        
        num_dato = 0
        while num_dato < constantes_clientes_frecuentes_excel.cantidad_total_variables_cliente.value-1: 
            if(datos[num_dato]!= None and datos[num_dato] != ""):
                clientes.cell(row=row, column=cant_var, value=datos[num_dato])
            cant_var += 1
            num_dato += 1

        excel_dataframe.save(path_excel)