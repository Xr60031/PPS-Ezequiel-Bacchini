from openpyxl import load_workbook
from Constantes.Facturacion.constantes_arrays import constantes_configuracion
from Constantes.Excel.constantes_excel import workbooks

class Facade_Configuracion():
    def __init__(self):
        self
    
    def guardar_configuracion_inicial(self, template_path, config, copy_path):
        excel_dataframe=load_workbook(template_path, data_only=False)
        datos_hoja=excel_dataframe[workbooks.datos.value]
        datos_hoja["A2"]=config[constantes_configuracion.pos_nombre_apellido.value]
        datos_hoja["B2"]=config[constantes_configuracion.pos_cuit.value]
        datos_hoja["C2"]=config[constantes_configuracion.pos_nombre_empresa.value]
        datos_hoja["D2"]=config[constantes_configuracion.pos_pventa.value]
        datos_hoja["E2"]=config[constantes_configuracion.pos_razon_social.value]
        datos_hoja["F2"]=config[constantes_configuracion.pos_domicilio.value]
        datos_hoja["G2"]=config[constantes_configuracion.pos_condicion_IVA.value]
        datos_hoja["H2"]=config[constantes_configuracion.pos_ingresos_brutos.value]
        datos_hoja["I2"]=config[constantes_configuracion.pos_fecha_inicio.value]
        datos_hoja["J2"]=config[constantes_configuracion.pos_iag.value]
        datos_hoja["K2"]=config[constantes_configuracion.pos_diag.value]
        datos_hoja["L2"]=config[constantes_configuracion.pos_alicuota.value]
        excel_dataframe.save(copy_path)
    
    def guardar_configuracion(self, path_excel, config):

        excel_dataframe=load_workbook(path_excel)
        
        datos_hoja=excel_dataframe[workbooks.datos.value]
        
        if config[constantes_configuracion.pos_nombre_apellido.value]:
            datos_hoja["A2"]=config[constantes_configuracion.pos_nombre_apellido.value]
        
        if config[constantes_configuracion.pos_cuit.value]:
            datos_hoja["B2"]=config[constantes_configuracion.pos_cuit.value]
        
        if config[constantes_configuracion.pos_nombre_empresa.value]:
            datos_hoja["C2"]=config[constantes_configuracion.pos_nombre_empresa.value]
        
        if config[constantes_configuracion.pos_pventa.value]:
            datos_hoja["D2"]=config[constantes_configuracion.pos_pventa.value]
        
        if config[constantes_configuracion.pos_razon_social.value]:
            datos_hoja["E2"]=config[constantes_configuracion.pos_razon_social.value]
        
        if config[constantes_configuracion.pos_domicilio.value]:
            datos_hoja["F2"]=config[constantes_configuracion.pos_domicilio.value]
        
        if config[constantes_configuracion.pos_condicion_IVA.value]:
            datos_hoja["G2"]=config[constantes_configuracion.pos_condicion_IVA.value]
        
        if config[constantes_configuracion.pos_ingresos_brutos.value]:
            datos_hoja["H2"]=config[constantes_configuracion.pos_ingresos_brutos.value]
        
        if config[constantes_configuracion.pos_fecha_inicio.value]:
            datos_hoja["I2"]=config[constantes_configuracion.pos_fecha_inicio.value]
        
        if config[constantes_configuracion.pos_iag.value]:
            datos_hoja["J2"]=config[constantes_configuracion.pos_iag.value]
        
        if config[constantes_configuracion.pos_diag.value]:
            datos_hoja["K2"]=config[constantes_configuracion.pos_diag.value]
        
        if config[constantes_configuracion.pos_alicuota.value]:
            datos_hoja["L2"]=config[constantes_configuracion.pos_alicuota.value]
        
        excel_dataframe.save(path_excel)