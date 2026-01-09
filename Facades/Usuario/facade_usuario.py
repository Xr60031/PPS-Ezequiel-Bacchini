from Funciones.ID.obtenerdor_IDs import Obtenedor_ID
from openpyxl import load_workbook
from Constantes.Facturacion.constantes_arrays import constantes_vendedor

from Constantes.Excel.constantes_excel import constantes_genericas_excel, constantes_posicion_datos_usuario_excel, workbooks
class Facade_usuario_manager():
    def __init__(self):
        self

    def armar_biblioteca_vendedor(self, datos_vendedor):
        obtenedorID = Obtenedor_ID()
        bibloteca_datos_vendedor = {
            "Nombre": datos_vendedor[constantes_vendedor.NOMBRE.value],
            "CUIT": datos_vendedor[constantes_vendedor.CUIT.value],
            "Nombre_Empresa": datos_vendedor[constantes_vendedor.NOMBRE_EMPRESA.value],
            "Punto_de_venta": datos_vendedor[constantes_vendedor.PUNTO_VENTA.value],
            "Razon_Social": datos_vendedor[constantes_vendedor.RAZON_SOCIAL.value],
            "Domicilio": datos_vendedor[constantes_vendedor.DOMICILIO.value],
            "Condicion_frente_al_IVA": datos_vendedor[constantes_vendedor.CONDICION_IVA_VENDEDOR.value],
            "Ingresos_Brutos": datos_vendedor[constantes_vendedor.INGRESOS_BRUTOS.value],
            "Fecha_Inicio": datos_vendedor[constantes_vendedor.FECHA_INICIO.value],
            "iag": datos_vendedor[constantes_vendedor.IMPUESTO_ADICIONAL_GLOBAL.value],
            "desc_iag": datos_vendedor[constantes_vendedor.DESCRIPCION_IAG.value],
            "alicuota": datos_vendedor[constantes_vendedor.ALICUOTA.value],
            "id_tributo_global": obtenedorID.obtener_ID_tributo(datos_vendedor[constantes_vendedor.IMPUESTO_ADICIONAL_GLOBAL.value])
        }

        return bibloteca_datos_vendedor
    
    def obtener_datos_usuario(self, file_name):
        excel_dataframe=load_workbook(file_name, data_only=True)
        datos_basicos_hoja=excel_dataframe[workbooks.datos.value]
        info_usuario = []

        datos_usuario=1

        #nombre
        #cuit
        #nombre empresa
        #punto venta
        #razon social
        #domicilio
        #condicion iva vendedor
        #ingresos brutos
        #fecha inicio
        #impuesto adicional global
        #descripcion iag
        #alicuota
        while datos_usuario < constantes_posicion_datos_usuario_excel.pos_Alicuota.value+1:
            info_usuario.append(datos_basicos_hoja.cell(row=constantes_genericas_excel.starting_row.value, column=datos_usuario).value)
            datos_usuario+=1

        excel_dataframe.close()

        return info_usuario