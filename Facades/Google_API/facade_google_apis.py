from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import googleapiclient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
import os

from Constantes.Flask.constantes_flask import constantes_de_flask
from Constantes.Excel.constantes_excel import constantes_genericas_excel, constantes_posicion_datos_factura_excel, workbooks
from Constantes.Google_API.constantes_google_api import constantes_google_api_

class Facade_API_google():
    def __init__(self):
        self.SERVICE_ACCOUNT_FILE_DRIVE = "facturadoreb-b5cf4202af91.json"
        self.SCOPES = ["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
        self.API_NAME = "drive"
        self.API_VERSION = "v3"

        self.credentials_drive = service_account.Credentials.from_service_account_file(self.SERVICE_ACCOUNT_FILE_DRIVE, scopes=self.SCOPES)
        # Crear el servicio de Google Drive
        self.drive_service = build("drive", "v3", credentials=self.credentials_drive)
        # Crear el servicio de Google Sheets
        self.sheets_service = build('sheets', 'v4', credentials=self.credentials_drive)
        # ID de la carpeta compartida en Google Drive
        self.FOLDER_ID = "1BW6PUmTcQoGPg7XyRRttas-1FolbTS-D"
    
    from openpyxl import load_workbook

    def __limpiar_excel(self, path):
        wb = load_workbook(path, data_only=False)
        
        for i in range(workbooks.cantidad_hojas.value):
            if(i==0):
                workbook = wb[workbooks.facturaA.value]
                column_start = constantes_posicion_datos_factura_excel.col_ultimo_dato_facturacion.value
            elif(i==1):
                workbook = wb[workbooks.facturaB.value]
                column_start = constantes_posicion_datos_factura_excel.col_ultimo_dato_facturacion.value
            elif(i==2):
                workbook = wb[workbooks.facturaC.value]
                column_start = constantes_posicion_datos_factura_excel.col_ultimo_dato_facturacion.value
            elif(i==3):
                workbook = wb[workbooks.xml.value]
                column_start = constantes_genericas_excel.starting_colvar.value
            elif(i==4):
                workbook = wb[workbooks.datos.value]
                column_start = constantes_genericas_excel.starting_colvar.value
            elif(i==5):
                workbook = wb[workbooks.items.value]
                column_start = constantes_genericas_excel.starting_colvar.value
            elif(i==6):
                workbook = wb[workbooks.historial.value]
                column_start = constantes_genericas_excel.starting_colvar.value
            elif(i==7):
                workbook = wb[workbooks.contactos_frecuentes.value]
                column_start = constantes_genericas_excel.starting_colvar.value

            for j in range(column_start, workbook.max_column + 1):
                col_letter = get_column_letter(j)
                if col_letter in workbook.column_dimensions:
                    del workbook.column_dimensions[col_letter]

            for row in workbook.iter_rows(min_row=constantes_genericas_excel.starting_row.value, max_row=workbook.max_row, min_col=constantes_genericas_excel.starting_colvar.value, max_col=workbook.max_column):
                workbook.row_dimensions[i].height = None
                for cell in row:
                    cell._style = None
            
            wb.save(path)

    def subir_excel(self, path_excel, excel_name):
        self.__limpiar_excel(path_excel)

        file_metadata = {
            "name": excel_name,
            "parents": self.FOLDER_ID,
            "mimeType": "application/vnd.google-apps.spreadsheet"
        }

        media = MediaFileUpload(
            path_excel,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )

        uploaded_file = self.drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        file_id = uploaded_file.get("id")

        spreadsheet = self.sheets_service.spreadsheets().get(spreadsheetId=file_id).execute()
        sheets = spreadsheet['sheets']

        requests = []

        hojas_objetivo = ["factura_a", "factura_b", "factura_c"]

        for sheet in sheets:
            sheet_title = sheet['properties']['title']
            sheet_id = sheet['properties']['sheetId']

            if sheet_title.strip().lower() in hojas_objetivo:
                requests.append({
                    "setDataValidation": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "startColumnIndex": constantes_posicion_datos_factura_excel.pos_Producto_Servicio.value - 1,
                            "endColumnIndex": constantes_posicion_datos_factura_excel.pos_Producto_Servicio.value
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_RANGE",
                                "values": [
                                    {"userEnteredValue": "=Items!$A$2:$A$101"}
                                ]
                            },
                            "showCustomUi": True
                        }
                    }
                })

        if requests:
            self.sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=file_id,
                body={"requests": requests}
            ).execute()

        self.drive_service.permissions().create(
            fileId=file_id,
            body={"role": "writer", "type": "anyone"}
        ).execute()

        file_id_embed = []
        file_id_embed.append(file_id)
        file_id_embed.append(f"https://docs.google.com/spreadsheets/d/{file_id}/edit?usp=sharing&rm=minimal")

        return file_id_embed
    
    def download_file_from_drive(self, service, file_id, destination_path):
        """
        Descarga un archivo desde Google Drive.

        :param service: Servicio autenticado de Google Drive.
        :param file_id: ID del archivo a descargar.
        :param destination_path: Ruta local donde guardar el archivo.
        """
        try:
            request = service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            with open(destination_path, 'wb') as file:
                downloader = googleapiclient.http.MediaIoBaseDownload(file, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    print(f"Descargando... {int(status.progress() * 100)}% completado.")
            print(f"Archivo descargado exitosamente en: {destination_path}")
        except Exception as e:
            print(f"Error al descargar el archivo: {e}")

    def delete_file_from_drive(self, service, file_id):
        """
        Elimina un archivo desde Google Drive.

        :param service: Servicio autenticado de Google Drive.
        :param file_id: ID del archivo a eliminar.
        """
        try:
            service.files().delete(fileId=file_id).execute()
            print(f"Archivo con ID {file_id} eliminado exitosamente.")
        except Exception as e:
            print(f"Error al eliminar el archivo: {e}")

    def delete_file_and_drive_after_download(self, file_path, file_ID):
        time.sleep(constantes_de_flask.sleep_time)
        self.delete_file_from_drive(self.drive_service, file_ID)
        os.remove(file_path)


    def delete_drive(self, file_ID):
        time.sleep(constantes_de_flask.sleep_time)
        self.delete_file_from_drive(self.drive_service, file_ID)

    def borrar_validacion_columna(self, file_name, columna = constantes_posicion_datos_factura_excel.pos_Producto_Servicio.value):
        excel_dataframe=load_workbook(file_name, data_only=True)

        for i in range(1, 3):
            if i == 1:
                hoja_facturacion=excel_dataframe[workbooks.facturaA.value]
            if i == 2:
                hoja_facturacion=excel_dataframe[workbooks.facturaB.value]
            if i == 3:
                hoja_facturacion=excel_dataframe[workbooks.facturaC.value]

            validaciones_a_eliminar = []
            for dv in hoja_facturacion.data_validations.dataValidation:
                for rango in dv.ranges:
                    if rango.min_col == columna and rango.max_col == columna:
                        validaciones_a_eliminar.append(dv)

            for dv in validaciones_a_eliminar:
                hoja_facturacion.data_validations.dataValidation.remove(dv)

        excel_dataframe.save(file_name)


    def descargar_excel(self, borrado, excelID, excelName):
        file_ID = excelID
        file_Name = excelName
        destination_path = f"uploads/{file_Name}"

        self.download_file_from_drive(self.drive_service, file_ID, destination_path)
        if(borrado == constantes_google_api_.si_borrado.value):
            self.delete_drive(file_ID)

        self.borrar_validacion_columna(destination_path)
        return destination_path