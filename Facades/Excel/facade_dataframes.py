from openpyxl import load_workbook
from Constantes.Excel.constantes_excel import constantes_xml_excel, workbooks
class Facade_dataframe():
    def __init__(self):
        self

    def obtener_dataframe_historial(self, file_name):
        excel_dataframe=load_workbook(file_name, data_only=True)
        datos_historial = []
        datos_historial.append(excel_dataframe)
        datos_historial.append(excel_dataframe[workbooks.historial.value])
        return datos_historial
    
    def obtener_xml(self, file_name):
        excel_dataframe=load_workbook(file_name, data_only=True)
        workbook = excel_dataframe[workbooks.xml.value]
        xml = workbook.cell(row=constantes_xml_excel.starting_row.value, column=constantes_xml_excel.starting_colvar.value).value
        excel_dataframe.save(file_name)
        return xml
    
    def guardar_xml(self, file_name, xml):
        excel_dataframe=load_workbook(file_name, data_only=True)
        workbook = excel_dataframe[workbooks.xml.value]
        workbook.cell(row=constantes_xml_excel.starting_row.value, column=constantes_xml_excel.starting_colvar.value, value=xml)
        excel_dataframe.save(file_name)
