from Funciones.Obtener_datos_facturacion.obtenedor_datos_facturacion import obtenedor_datos_facturacion
from openpyxl import load_workbook
from Constantes.Excel.constantes_excel import workbooks, constantes_genericas_excel
#FM = Facturador Multiple
class Obtenedor_FM(obtenedor_datos_facturacion):
    def __init__(self):
        super().__init__()

    def obtener_worksheet_facturacion(self, filename):
        excel_dataframe=load_workbook(filename, data_only=True)
        worksheet = None
        i = 0
        while worksheet == None:
            factura_hoja=excel_dataframe[workbooks.facturaA.value]
            if(factura_hoja.cell(row=constantes_genericas_excel.starting_row.value+i, column=constantes_genericas_excel.starting_colvar.value).value):
                worksheet = workbooks.facturaA.value
            factura_hoja=excel_dataframe[workbooks.facturaB.value]
            if(factura_hoja.cell(row=constantes_genericas_excel.starting_row.value+i, column=constantes_genericas_excel.starting_colvar.value).value):
                worksheet = workbooks.facturaB.value
            factura_hoja=excel_dataframe[workbooks.facturaC.value]
            if(factura_hoja.cell(row=constantes_genericas_excel.starting_row.value+i, column=constantes_genericas_excel.starting_colvar.value).value):
                worksheet = workbooks.facturaC.value
            
            i+=1
        
        return worksheet

    def obtener_datos_facturacion(self, data_source, datos_factura_manager, datos_usuario):
        bibliotecas_facturas = []
        worksheet = self.obtener_worksheet_facturacion(data_source)
        biblioteca_datos = datos_factura_manager.obtener_datos_factura_multiple(data_source, worksheet)
        for i in range(len(biblioteca_datos)):
            datos = self.calcular_datos_items_tributos(biblioteca_datos[i]["items"], biblioteca_datos[i]["tributos"], biblioteca_datos[i]["datos_factura"], datos_usuario)
            bibliotecas_facturas.append(self.armar_biblioteca_factura(datos))
        return bibliotecas_facturas
